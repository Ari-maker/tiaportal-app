from flask import Flask, render_template, request, url_for, redirect, jsonify
from flaskwebgui import FlaskUI # import FlaskUI
from email.mime.text import MIMEText
import smtplib
from email.message import EmailMessage
import openpyxl
import json
from treelib import Node, Tree
app = Flask(__name__)

  
#from openness import startTiaPortal
#startTiaPortal()
#from openness import startProcess
from tkinter.filedialog import askopenfilename


#_type = '6SL3210-1KE11-8AF2/4.7.13';
#_name = 'Device_1';

#tree = Tree()
#tree.create_node("Harry", "harry")  # root node
#tree.create_node("Jane", "jane", parent="harry")
#tree.create_node("Bill", "bill", parent="harry")
#tree.create_node("Diane", "diane", parent="jane")
#tree.create_node("Mary", "mary", parent="diane")
#tree.create_node("Mark", "mark", parent="jane")
#tree.show()


_typeArr = []
_nameArr = []
_consoleArr = []

myproject = ''

mypath = ''

counter = 2 # 3

dllpath = ''

libpath = ''

interface = True

tia = []
project = []
dlist = []

tree = Tree()

tree.create_node("Parent", "Parent")  # root node
tree.create_node("DrivesData [DB]", "DrivesData [DB]", parent="Parent")
tree.create_node("Drives [OB]", "Drives [OB]", parent="Parent")

  
@app.route("/openproject/", methods=["GET"])
def openproject():
    print('project path')
    global mypath
    mypath = askopenfilename(filetypes=(("Project file", "*.ap18"),("All files", "*.*") ))


    return jsonify({'project':mypath})


@app.route("/selectdll/", methods=["GET"])
def selectdll():
    print('dll path')
    global dllpath
    dllpath = askopenfilename(filetypes=(("Siemens.Engineering file", "*.dll"),("All files", "*.*") ))


    return jsonify({'dll':dllpath})


@app.route("/selectlib/", methods=["GET"])
def selectlib():
    print('lib path')
    global libpath
    libpath = askopenfilename(filetypes=(("al file", "*.al18"),("All files", "*.*") ))


    return jsonify({'lib':libpath})


@app.route("/delete/", methods=["POST"])
def delete():
 
    name = request.form['name']
    name = name.capitalize()
    index = _nameArr.index(name)
    _nameArr.remove(name)
    _typeArr.pop(index)

    removeDIR = "Inst"+name+" [DB]"

    tree.remove_node(removeDIR)

    return jsonify({'type':_typeArr, 'name':_nameArr})


@app.route("/load/")
def load():


    global mypath 
    global dllpath
    global libpath

    # Opening JSON file
    with open('paths.json', 'r') as openfile:
       # Reading from json file
        json_object = json.load(openfile)
        mypath = json_object['project']
        dllpath = json_object['dll']
        libpath = json_object['lib']

   
    return jsonify({'project':mypath, 'dll':dllpath,'lib':libpath})


@app.route("/getDevices/")
def getDevices():
        return jsonify({'type':_typeArr, 'name':_nameArr})

@app.route("/addDevice/", methods=["POST"])
def add():

    if request.method == "POST":
            _type =  request.form['type']
            _name =  request.form['name']
 

    if _type == '' or _name == '':
        return 400
    

    _name = _name.replace(" ", "")
    _name = _name.capitalize()

    global _typeArr
    global _nameArr
    if _name in _nameArr:
       return 400
 
    _typeArr.append(_type)
    _nameArr.append(_name)
 
    print(_typeArr)
    
    print(_nameArr)

    tree.create_node("Inst"+_name+" [DB]", "Inst"+_name+" [DB]", parent="Parent")

    return jsonify({'type':_typeArr, 'name':_nameArr})
 
 

@app.route("/")
def index(): 
    # resetoidaan
    global myproject
    #myproject = ''
    global mypath 
    #mypath = ''
    global _typeArr
    #_typeArr = []
    global _nameArr
    #_nameArr = []
    global interface
    interface = True
   

    return render_template("index.html")

def getXML():

    print('XML')

    global _nameArr
    global counter

    from openness import writeXML
    writeXML(_nameArr, counter)

    return render_template("index.html")


@app.route("/add")
def addFunc():

    name = request.args.get('name')



    parentArr = json.loads(tree.to_json(with_data=False))


    
    def treeToHtml(array, content, nest):
      
        for node in array:

            if isinstance(node, dict):
          
                currentkey = ''
                for key in node.keys(): 
                    currentkey = key
                    
                content += '<li id="'+currentkey+'" role="treeitem" aria-expanded="false" aria-selected="false" draggable="true" ondragstart="drag(event)"><span>'+currentkey+'</span><ul role="group">'

                for file in node[currentkey]["children"]:
                 
                    if isinstance(file, dict):
                        for key in file.keys(): 
                        
                            content = treeToHtml(node[currentkey]["children"], content, True)
                   
                    else:
                        content += '<li id="'+file+'" role="treeitem" aria-selected="false" class="doc" draggable="true" ondragstart="drag(event)">'+file+'</li>'
                

                content += '</ul></li>'

            elif nest is False:   
                content += '<li id="'+node+'" role="treeitem" aria-selected="false" class="doc" draggable="true" ondragstart="drag(event)">'+node+'</li>'


        return content
    
    content = ''
    html = treeToHtml(parentArr["Parent"]["children"], content, False)

    tree.create_node(name, name, parent="Parent")

    html += '<li id="'+name+'" role="treeitem" aria-expanded="false" aria-selected="false" draggable="true" ondragstart="drag(event)"><span>'+name+'</span><ul role="group"></ul></li>'

    return render_template('device.html', mylist=html)


@app.route("/tia/")
def tiafunc():
    if len(tia) != 0:
        return jsonify({'tia':1, 'dlist':dlist})

    return jsonify({'tia':0,})

@app.route("/tiaportal/", methods=["POST"])
def initTia():
    global _typeArr
    global _nameArr
    global mypath
    global _consoleArr
    global dllpath
    global libpath
    global interface
   
    print(_typeArr)
    print(_nameArr)
    print(mypath)
    

    if not _typeArr: 
        return 400
    
    if not _nameArr:
        return 400
   
    if not dllpath:
        return 400
    
    if not libpath:
       return 400
    
    if not tia:
        return 400
    
    if not project:
        return 400
    
    getXML()

    directory = json.loads(tree.to_json(with_data=False))

    from openness import startProcess
    startProcess(_typeArr, _nameArr, _consoleArr, dllpath, libpath, tia[0], project[0], directory)

    #return redirect('/')
    return render_template("device.html")


@app.route("/initPortal/", methods=["POST"])
def tiaportal():
    
    global _typeArr
    global _nameArr
    global mypath
    global _consoleArr
    global dllpath
    global libpath
    global interface

    _consoleArr = []

    print(_typeArr)
    print(_nameArr)
    print(mypath)
    
    if not mypath:
        return 400
   
    if not dllpath:
        return 400
    
    if not libpath:
       return 400
    
 
    
    getXML()

    #from openness import startProcess
    #startProcess(_typeArr, _nameArr, mypath, _consoleArr, dllpath, libpath, tia)


    from openness import initProcess
    initProcess(_consoleArr, mypath,dllpath,interface, tia, project, dlist)
     
    #return redirect('/')

    if len(tia) != 0:
        return jsonify({'tia':1, 'dlist':dlist})

    return jsonify({'tia':0})

@app.route("/console/")
def console():
     return render_template("console.html")

@app.route("/device/")
def device():

    # parse x:
    parentArr = json.loads(tree.to_json(with_data=False))
  
    def treeToHtml(array, content, nest):
      
        for node in array:

            if isinstance(node, dict):
          
                currentkey = ''
                for key in node.keys(): 
                    currentkey = key
                    
                content += '<li id="'+currentkey+'" role="treeitem" aria-expanded="false" aria-selected="false" draggable="true" ondragstart="drag(event)"><span>'+currentkey+'</span><ul role="group">'

                for file in node[currentkey]["children"]:
                 
                    if isinstance(file, dict):
                        for key in file.keys(): 
                        
                            content = treeToHtml(node[currentkey]["children"], content, True)
                   
                    else:
                        content += '<li id="'+file+'" role="treeitem" aria-selected="false" class="doc" draggable="true" ondragstart="drag(event)">'+file+'</li>'
                

                content += '</ul></li>'

            elif nest is False:   

                if "[" in node:
                    content += '<li id="'+node+'" role="treeitem" aria-selected="false" class="doc" draggable="true" ondragstart="drag(event)">'+node+'</li>'
                else:  
                    content += '<li id="'+node+'" role="treeitem" aria-expanded="false" aria-selected="false" draggable="true" ondragstart="drag(event)"><span>'+node+'</span><ul role="group"></ul></li>'  


        return content
    
    content = ''
    html = treeToHtml(parentArr["Parent"]["children"], content, False)
    
    return render_template('device.html', mylist=html)



@app.route("/tiaportalExcel/", methods=["POST"])
def tiaportalExcel():
    
    global _typeArr
    global _nameArr
    global mypath
    global _consoleArr
    global dllpath
    global libpath
    global interface
    _consoleArr = []

    print(_typeArr)
    print(_nameArr)
    print(mypath)
    
    if not mypath:
        return 400

    if not _typeArr: 
        return 400
    
    if not _nameArr:
        return 400
   
    if not dllpath:
        return 400
    
    if not libpath:
       return 400
    
    getXML()

    from openness import startProcessExcel
    startProcessExcel(_typeArr, _nameArr, mypath, _consoleArr, dllpath, libpath, interface)
    

    return render_template("index.html")


@app.route("/importExcel/")
def importExcel():

    excelPath = askopenfilename(filetypes=(("Excel file", "*.xlsx"),("All files", "*.*") ))

    if excelPath is None or excelPath == "":
        return render_template("index.html")

    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(excelPath)
    
    # Define variable to read sheet
    dataframe1 = dataframe.active
    
    foundPaths = True
    foundDevices = False
    pathsRow = -1

    # Iterate the loop to read the cell values
    for row in range(0, dataframe1.max_row):
        index = 0

        for col in dataframe1.iter_cols(1, dataframe1.max_column):
        

            # paths
            if col[row].value != None and foundPaths == True:

                pathsRow = row+1

                if(index == 0):
                    global mypath
                    mypath = col[pathsRow].value
                if(index == 1):
                    global dllpath
                    dllpath = col[pathsRow].value
                if(index == 2):
                    global libpath
                    libpath = col[pathsRow].value
                    foundPaths = False
                    foundDevices = True

       # devices
            elif col[row].value != None and foundDevices == True and row > (pathsRow+1):
              
                if(index == 0):
                    _nameArr.append(col[row].value)
                if(index == 1):
                    _typeArr.append(col[row].value)
              

            index = index + 1

    return jsonify({'type':_typeArr, 'name':_nameArr})


@app.route("/interface/")
def interfaceUser():

    global interface

    if interface is True:
        interface = False
    else:
        interface = True

    return jsonify({'interface':str(interface)})


@app.route("/removeDirectory/", methods=["POST"])
def removeDir():
    _dir = request.form['folder']
   
    tree.remove_node(_dir)

    return render_template('index.html') 


@app.route("/addDirectory/", methods=["POST"])
def addDir():
    _dir = request.form['foldername']
    tree.create_node(_dir, _dir, parent="Parent")

    return render_template('device.html') 


@app.route("/updateFile/", methods=["POST"])
def updateDIR():
        _folder = request.form['folder']
        _file = request.form['file']

        print(_folder)
        print(_file)

        global tree

        if _folder == "":
            tree.move_node(_file, "Parent") 
        else:
            tree.move_node(_file, _folder) 
      
        print("tree: "+ tree.to_json(with_data=False))

        return render_template('device.html') 



@app.route("/saveFile/", methods=["POST"])
def saveFile():
    _project = request.form['project']
    _dll = request.form['dll']
    _lib = request.form['lib']

    filedata = {
    "project": _project,
    "dll": _dll,
    "lib": _lib
    }

    # Serializing json
    json_object = json.dumps(filedata, indent=3)

    # Writing to sample.json
    with open("paths.json", "w") as outfile:
        outfile.write(json_object)

    return jsonify({'file':str(filedata)})  


@app.route("/getConsoleData/")
def consoleData():
   
     global _consoleArr

     print(_consoleArr)

     return jsonify({'consoleArr':str(_consoleArr)})
    

if __name__ == "__main__":
    #app.run(debug=True)
    #ui.run()
    FlaskUI(app=app, width=1280 , height=800, server="flask").run()
     #ui = FlaskUI(app, width=500, height=500) # add app and parameters